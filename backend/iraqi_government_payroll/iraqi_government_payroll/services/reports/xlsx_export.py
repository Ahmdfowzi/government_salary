# Copyright (c) 2026, Iraqi Government Payroll
"""Render report data to an .xlsx workbook with openpyxl. Pure — no Frappe.

Takes rows the report aggregators already produced plus a column spec, and lays
them out: a bold frozen Arabic header row (RTL sheet), the data rows, and an
optional bold totals row. No aggregation here. Booleans render as نعم/لا and
lists (e.g. the bank-transfer `missing` reasons) join with a comma so flagged
incomplete rows are exported in full, never dropped.
"""

import io

from openpyxl import Workbook
from openpyxl.styles import Font

_TOTAL_LABEL = "الإجمالي"


def _cell(value):
	if value is None:
		return ""
	if isinstance(value, bool):
		return "نعم" if value else "لا"
	if isinstance(value, (list, tuple)):
		return "، ".join(str(x) for x in value)
	return value


def build_workbook(title, columns, rows, totals=None):
	"""Return .xlsx bytes for `rows` laid out by `columns` [(key, header), ...].

	`totals` is an optional {column_key: value} dict rendered as a final bold row
	(label in the first column).
	"""
	wb = Workbook()
	ws = wb.active
	ws.title = (title or "Report")[:31]
	ws.sheet_view.rightToLeft = True
	bold = Font(bold=True)

	ws.append([header for _key, header in columns])
	for cell in ws[1]:
		cell.font = bold
	ws.freeze_panes = "A2"

	for row in rows:
		ws.append([_cell(row.get(key)) for key, _header in columns])

	if totals:
		line = []
		for idx, (key, _header) in enumerate(columns):
			if idx == 0:
				line.append(_TOTAL_LABEL)
			elif key in totals:
				line.append(_cell(totals[key]))
			else:
				line.append("")
		ws.append(line)
		for cell in ws[ws.max_row]:
			cell.font = bold

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()
