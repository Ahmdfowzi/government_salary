# Copyright (c) 2026, Iraqi Government Payroll
"""Live bench smoke checks for the Phase 3 governance / locking workflow.

Run each as a single process via `bench execute` (NOT `bench console`):

    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.governance
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.locking
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.api
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.create
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.reports
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.pension_report
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.bank_transfer
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.excel_export
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.pdf_export
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.accounting_journal
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.security
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.grade_validation
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.employee_profile_api
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.payroll_slip
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.data_integrity
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.transaction_forms
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.uat_demo_cycle
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.financial_wiring
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.family_dependents
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.async_payroll_run
    bench --site payroll.localhost execute iraqi_government_payroll.smoke.checks.demo

Why `bench execute`: each check runs in one process with a proper Frappe context,
so the first failed assertion raises, aborts the run with a non-zero exit code,
and the "PASSED" line — the last statement of each function — is unreachable once
anything has failed. The old `bench console < script.py` harness executed line by
line: it swallowed per-line exceptions (printing a false "PASSED") and tripped a
document-lock race during calculate. These functions only build/read records and
drive the governance API; they contain no calculation, governance or DocType logic.
"""

import frappe


def _blocked(fn):
	"""Return True if calling fn() raises (i.e. the action was blocked)."""
	try:
		fn()
		return False
	except Exception:
		return True


def governance():
	"""Phase 3 M1 — approval path + protection rules + M5 audit-event trail."""
	# Reuse / create a valid employee profile (Bachelor g7s1 -> 0 calc errors)
	if not frappe.db.exists("Government Employee Payroll Profile", {"employee_number": "E1"}):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": "E1", "employee_name": "Smoke Test",
			"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
			"current_stage": 1, "qualification": "Bachelor", "status": "Active",
		}).insert()

	if not frappe.db.exists("Payroll Period", {"year": 2020, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2020, "month": 6,
						"start_date": "2020-06-01", "end_date": "2020-06-30", "status": "Open"}).insert()
	period = frappe.get_value("Payroll Period", {"year": 2020, "month": 6}, "name")

	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "All"}).insert()
	print("after insert        :", run.workflow_state)            # Draft

	run.calculate_run(); run.reload()
	print("after calculate     :", run.workflow_state, "| run_status:", run.run_status,
		  "| errors:", run.error_count, "| calculated_by:", run.calculated_by)
	assert run.workflow_state == "Calculated", run.workflow_state

	run.submit_for_review(); run.reload()
	print("after submit_review :", run.workflow_state, "| reviewed_by:", run.reviewed_by)
	assert run.workflow_state == "Under Review", run.workflow_state

	run.approve_run(); run.reload()
	print("after approve       :", run.workflow_state, "| approved_by:", run.approved_by)
	assert run.workflow_state == "Approved", run.workflow_state

	# Protection rule: cannot recalculate after approval
	recalc_blocked = _blocked(lambda: run.calculate_run())
	print("recalc blocked      :", recalc_blocked)
	assert recalc_blocked, "recalculation allowed after approval"

	run.submit_run(); run.reload()
	print("after submit        :", run.workflow_state, "| submitted_by:", run.submitted_by)
	assert run.workflow_state == "Submitted", run.workflow_state

	# Protection rule: cannot delete a submitted run
	delete_blocked = _blocked(lambda: frappe.delete_doc("Payroll Run", run.name))
	print("delete blocked      :", delete_blocked)
	assert delete_blocked, "delete allowed on submitted run"

	# M5: every transition left exactly one immutable governance event, in order.
	events = frappe.get_all("Payroll Run Governance Event",
							filters={"payroll_run": run.name},
							fields=["action", "from_state", "to_state"],
							order_by="creation asc")
	actions = [e["action"] for e in events]
	print("governance events   :", actions)
	assert actions == ["calculate", "submit_for_review", "approve", "submit"], actions

	frappe.db.commit()
	print("\nGOVERNANCE SMOKE TEST PASSED")


def locking():
	"""Phase 3 M3 / M3.1 — lock, historical integrity, retroactive protection."""
	from iraqi_government_payroll.services.lifecycle import lifecycle_service as lc
	from iraqi_government_payroll.services.historical import history_service as hist

	EMP = "LOCK1"
	if not frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": EMP, "employee_name": "Lock Test",
			"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
			"current_stage": 1, "qualification": "Bachelor", "status": "Active",
			"employment_status": "Active",
		}).insert()

	period = frappe.db.get_value("Payroll Period", {"year": 2026, "month": 1})
	if not period:
		period = frappe.get_doc({"doctype": "Payroll Period", "year": 2026, "month": 1,
								 "start_date": "2026-01-01", "end_date": "2026-01-31",
								 "status": "Open"}).insert().name

	# Run -> approve -> submit -> lock
	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "Employee",
						  "scope_reference": EMP}).insert()
	run.calculate_run(); run.reload()
	# The run builds DRAFT slips; submitting one fires Salary Slip.on_submit, which
	# writes the immutable Payroll Calculation Snapshot the historical-integrity
	# checks below read back. (Governance run state is independent of slip docstatus.)
	slip = frappe.db.get_value(
		"Salary Slip", {"payroll_run": run.name, "employee_profile": EMP}, "name")
	assert slip, "run did not produce a Salary Slip for the employee"
	slip_doc = frappe.get_doc("Salary Slip", slip)
	if slip_doc.docstatus == 0:
		slip_doc.submit()
	run.submit_for_review(); run.reload()
	run.approve_run(); run.reload()
	run.submit_run(); run.reload()
	run.lock_run(); run.reload()
	print("workflow_state      :", run.workflow_state, "| locked_by:", run.locked_by)
	assert run.workflow_state == "Locked", run.workflow_state
	assert run.is_run_locked() and run.is_historical_period()

	net_before = hist.get_payroll_snapshot(run.name, EMP)["net_amount"]
	print("historical net      :", net_before)

	# Post-lock (future-dated) change is allowed; profile mutated to simulate promotion
	lc.transfer_employee(EMP, "2026-06-01", to_entity=None, reason="post-lock transfer")
	prof = frappe.get_doc("Government Employee Payroll Profile", EMP)
	prof.current_stage = 5
	prof.save()
	print("profile changed     : stage ->",
		  frappe.db.get_value("Government Employee Payroll Profile", EMP, "current_stage"))

	# Retroactive change INTO the locked period must be blocked
	retro_blocked = _blocked(lambda: lc.transfer_employee(EMP, "2026-01-15", reason="retroactive"))
	print("retroactive blocked :", retro_blocked)
	assert retro_blocked, "retroactive transfer into locked period was allowed"

	# Historical snapshot unchanged
	net_after = hist.get_payroll_snapshot(run.name, EMP)["net_amount"]
	print("historical net after:", net_after)
	assert net_after == net_before, "locked payroll snapshot changed!"

	# Recalculation must be blocked
	recalc_blocked = _blocked(lambda: run.calculate_run())
	print("recalc blocked      :", recalc_blocked)
	assert recalc_blocked, "recalculation allowed after lock"

	# Deletion must be blocked
	delete_blocked = _blocked(lambda: frappe.delete_doc("Payroll Run", run.name))
	print("delete blocked      :", delete_blocked)
	assert delete_blocked, "delete allowed after lock"

	frappe.db.commit()
	print("\nPAYROLL LOCKING SMOKE TEST PASSED")


def api():
	"""Phase 3 M6 — the governance REST surface (api.payroll_api) end to end.

	Drives a run through get_run_governance + run_governance_action and checks that
	state, available actions and the M5 audit trail flow through the API exactly as
	through the controller. Runs as Administrator (System Manager bypasses roles).
	"""
	from iraqi_government_payroll.api import payroll_api as papi

	if not frappe.db.exists("Government Employee Payroll Profile", {"employee_number": "E1"}):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": "E1", "employee_name": "Smoke Test",
			"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
			"current_stage": 1, "qualification": "Bachelor", "status": "Active",
		}).insert()
	if not frappe.db.exists("Payroll Period", {"year": 2020, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2020, "month": 6,
						"start_date": "2020-06-01", "end_date": "2020-06-30", "status": "Open"}).insert()
	period = frappe.get_value("Payroll Period", {"year": 2020, "month": 6}, "name")

	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "All"}).insert()

	g0 = papi.get_run_governance(run.name)
	print("initial state       :", g0["workflow_state"], "| actions:", g0["allowed_actions"])
	assert g0["workflow_state"] == "Draft", g0["workflow_state"]
	assert "calculate" in g0["allowed_actions"]

	r1 = papi.run_governance_action(run.name, "calculate")
	print("after calculate     :", r1["workflow_state"], "| actions:", r1["allowed_actions"])
	assert r1["workflow_state"] == "Calculated", r1["workflow_state"]
	assert "submit_for_review" in r1["allowed_actions"]

	papi.run_governance_action(run.name, "submit_for_review")
	papi.run_governance_action(run.name, "approve")
	r2 = papi.run_governance_action(run.name, "submit")
	print("after submit        :", r2["workflow_state"], "| actions:", r2["allowed_actions"])
	assert r2["workflow_state"] == "Submitted", r2["workflow_state"]

	# Unknown action is rejected before any dispatch
	bad = _blocked(lambda: papi.run_governance_action(run.name, "nuke"))
	print("unknown rejected    :", bad)
	assert bad, "unknown action was not rejected"

	# The same immutable M5 audit trail is visible through the read endpoint
	g1 = papi.get_run_governance(run.name)
	logged = [e["action"] for e in g1["events"]]
	print("events via API      :", logged)
	assert logged == ["calculate", "submit_for_review", "approve", "submit"], logged

	frappe.db.commit()
	print("\nGOVERNANCE API SMOKE TEST PASSED")


def create():
	"""Phase 3 M8 — the create_payroll_run endpoint: validation + duplicate guard.

	Uses a dedicated period (2019/12) that no other check touches, and rolls back
	the created run at the end so the check is re-runnable and never collides with
	the duplicate guard on a later run.
	"""
	from iraqi_government_payroll.api import payroll_api as papi

	if not frappe.db.exists("Payroll Period", {"year": 2019, "month": 12}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2019, "month": 12,
						"start_date": "2019-12-01", "end_date": "2019-12-31",
						"status": "Open"}).insert()
		frappe.db.commit()                      # keep the period; only the run is rolled back
	period = frappe.get_value("Payroll Period", {"year": 2019, "month": 12}, "name")

	r = papi.create_payroll_run(period, "IRAQ-2015", "All", None)
	print("created run          :", r["name"], "|", r["workflow_state"], "| actions:", r["allowed_actions"])
	assert r["workflow_state"] == "Draft", r["workflow_state"]
	assert "calculate" in r["allowed_actions"]

	# Duplicate (same period + rule_set + scope, not cancelled) is blocked
	dup_blocked = _blocked(lambda: papi.create_payroll_run(period, "IRAQ-2015", "All", None))
	print("duplicate blocked    :", dup_blocked)
	assert dup_blocked, "duplicate active run was allowed"

	# Invalid inputs are rejected
	bad_period = _blocked(lambda: papi.create_payroll_run("NO-SUCH-PERIOD", "IRAQ-2015", "All", None))
	print("bad period rejected  :", bad_period)
	assert bad_period, "non-existent period was accepted"

	missing_ref = _blocked(lambda: papi.create_payroll_run(period, "IRAQ-2015", "Employee", None))
	print("missing ref rejected :", missing_ref)
	assert missing_ref, "Employee scope without a reference was accepted"

	frappe.db.rollback()                        # discard the created run; re-runnable
	print("\nPAYROLL RUN CREATE SMOKE TEST PASSED")


def reports():
	"""Phase 4 M10 — the read-only report endpoints reconcile to the Salary Slip.

	Builds a one-employee run (dedicated employee + period), calculates it (draft
	slips), reads all five reports and checks every total ties back to the slip,
	then rolls back so the check is re-runnable.
	"""
	from iraqi_government_payroll.api import reports_api as rep

	EMP = "REP1"
	if not frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": EMP, "employee_name": "Report Test",
			"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
			"current_stage": 1, "qualification": "Bachelor", "status": "Active",
			"employment_status": "Active",
		}).insert()
		frappe.db.commit()                      # keep the profile; only the run is rolled back

	if not frappe.db.exists("Payroll Period", {"year": 2018, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2018, "month": 6,
						"start_date": "2018-06-01", "end_date": "2018-06-30",
						"status": "Open"}).insert()
		frappe.db.commit()
	period = frappe.get_value("Payroll Period", {"year": 2018, "month": 6}, "name")

	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "Employee",
						  "scope_reference": EMP}).insert()
	run.calculate_run(); run.reload()

	summary = rep.run_summary(run.name)
	emp = rep.employee_register(run.name)
	allow = rep.allowances_register(run.name)
	ded = rep.deductions_register(run.name)
	tax = rep.tax_register(run.name)
	print("summary             :", summary)
	print("employee rows       :", emp["rows"])

	assert summary["employees"] == 1, summary
	assert len(emp["rows"]) == 1
	row = emp["rows"][0]
	# reconcile registers back to the slip totals (the single source)
	assert emp["totals"]["net"] == summary["total_net"], "net mismatch"
	assert allow["grand_total"] == emp["totals"]["allowances"], "allowances mismatch"
	assert ded["grand_total"] == emp["totals"]["deductions"], "deductions mismatch"
	assert tax["total_tax"] <= ded["grand_total"], "tax exceeds deductions"
	assert row["basic"] + row["allowances"] == summary["total_earnings"], "earnings mismatch"
	print("reconciled          : net/allowances/deductions/tax all tie to the slip")

	frappe.db.rollback()                        # discard the run; re-runnable
	print("\nPAYROLL REPORTS SMOKE TEST PASSED")


def pension_report():
	"""Phase 4 M11 — the Retirement Pension Register reads stored Pension
	Calculation values and reconciles. Inserts one sample record (no retirement
	calc here — figures are supplied as already-computed), reads the register,
	checks reconciliation, then rolls back.
	"""
	from iraqi_government_payroll.api import reports_api as rep

	if not frappe.db.exists("Government Employee Payroll Profile", "PEN1"):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": "PEN1", "employee_name": "Pension Test",
			"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
			"current_stage": 1, "qualification": "Bachelor", "status": "Active",
			"employment_status": "Retired",
		}).insert()
		frappe.db.commit()

	# Stored, already-computed pension figures (gross = approved + cert + col;
	# net = gross - tax - other). NOT computed here.
	approved, cert, col, tax, other, eos = 909000, 90900, 227250, 60000, 0, 12360000
	gross = approved + cert + col
	net = gross - tax - other
	frappe.get_doc({
		"doctype": "Pension Calculation", "employee_profile": "PEN1",
		"employee_name": "Pension Test", "rule_set": "IRAQ-2015",
		"calculation_date": "2026-06-15", "period_date": "2026-06-01", "status": "Calculated",
		"service_years": 36, "average_36_months": 1010000, "accrual_rate": 2.5,
		"approved_pension": approved, "certificate_allowance": cert, "cost_of_living": col,
		"gross_pension": gross, "monthly_tax": tax, "other_deductions": other,
		"net_pension": net, "end_of_service_bonus": eos,
	}).insert()

	reg = rep.pension_register("2026-06-01", "2026-06-30", None)
	print("pension rows        :", reg["rows"])
	print("pension totals      :", reg["totals"])
	assert reg["count"] == 1, reg["count"]
	row = reg["rows"][0]
	assert row["qualification"] == "Bachelor", row["qualification"]
	assert row["net_pension"] == net
	# reconciliation against stored figures (no recompute)
	assert row["gross_pension"] == row["approved_pension"] + row["certificate_allowance"] \
		+ row["cost_of_living"], "gross mismatch"
	assert row["net_pension"] == row["gross_pension"] - row["monthly_tax"] \
		- row["other_deductions"], "net mismatch"
	assert reg["totals"]["net_pension"] == net
	print("reconciled          : gross/net tie to stored pension figures")

	frappe.db.rollback()                        # discard the record; re-runnable
	print("\nRETIREMENT PENSION REGISTER SMOKE TEST PASSED")


def bank_transfer():
	"""Phase 4 M12 — Bank Transfer Export: net from the slip, bank from the profile,
	missing bank data flagged (never skipped). One employee has bank details, one
	does not; both appear. Rolls back so it is re-runnable.
	"""
	from iraqi_government_payroll.api import reports_api as rep

	# BANKED employee (has an account) and UNBANKED employee (no account/iban).
	for emp, name, acct in (("BANK1", "Banked", "1122334455"), ("BANK2", "Unbanked", None)):
		if not frappe.db.exists("Government Employee Payroll Profile", emp):
			doc = {
				"doctype": "Government Employee Payroll Profile",
				"employee_number": emp, "employee_name": name,
				"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
				"current_stage": 1, "qualification": "Bachelor", "status": "Active",
				"employment_status": "Active",
			}
			if acct:
				doc["bank_account"] = acct
				doc["bank_name"] = "Rafidain"
			frappe.get_doc(doc).insert()
			frappe.db.commit()

	if not frappe.db.exists("Payroll Period", {"year": 2017, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2017, "month": 6,
						"start_date": "2017-06-01", "end_date": "2017-06-30",
						"status": "Open"}).insert()
		frappe.db.commit()
	period = frappe.get_value("Payroll Period", {"year": 2017, "month": 6}, "name")

	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "All"}).insert()
	run.calculate_run(); run.reload()

	bt = rep.bank_transfer(run.name)
	print("bank rows           :", bt["rows"])
	print("incomplete_count    :", bt["incomplete_count"], "| total_net:", bt["total_net"])

	by = {r["employee_profile"]: r for r in bt["rows"]}
	assert "BANK1" in by and "BANK2" in by, "an employee row was skipped"   # nothing skipped
	assert by["BANK1"]["bank_complete"] is True, by["BANK1"]
	assert by["BANK1"]["bank_account"] == "1122334455"
	assert by["BANK2"]["bank_complete"] is False, by["BANK2"]
	assert "bank_account" in by["BANK2"]["missing"], by["BANK2"]["missing"]
	# net is read from the slip (positive for the banked active employee)
	assert by["BANK1"]["net"] > 0
	# total_net is the plain sum of slip nets (no recompute)
	assert bt["total_net"] == sum(r["net"] for r in bt["rows"])
	print("flagged             : BANK2 missing bank_account, included not skipped")

	frappe.db.rollback()                        # discard the run; re-runnable
	print("\nBANK TRANSFER SMOKE TEST PASSED")


def excel_export():
	"""Phase 4 M13 — Excel export renders a valid workbook for every report by
	reusing the report aggregators. Builds a run + a pension record, renders each
	of the 7 reports to xlsx, reopens one and checks the header, then rolls back.
	"""
	import io
	from openpyxl import load_workbook
	from iraqi_government_payroll.api import reports_api as rep

	EMP = "XLS1"
	if not frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": EMP, "employee_name": "Excel Test",
			"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
			"current_stage": 1, "qualification": "Bachelor", "status": "Active",
			"employment_status": "Active", "bank_account": "55667788",
		}).insert()
		frappe.db.commit()
	if not frappe.db.exists("Payroll Period", {"year": 2016, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2016, "month": 6,
						"start_date": "2016-06-01", "end_date": "2016-06-30",
						"status": "Open"}).insert()
		frappe.db.commit()
	period = frappe.get_value("Payroll Period", {"year": 2016, "month": 6}, "name")

	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "Employee",
						  "scope_reference": EMP}).insert()
	run.calculate_run(); run.reload()

	run_reports = ["run_summary", "employee_register", "allowances_register",
				   "deductions_register", "tax_register", "bank_transfer"]
	for report in run_reports:
		content = rep.render_report_xlsx(report, run=run.name)
		assert content[:2] == b"PK", f"{report}: not an xlsx"
		assert len(content) > 200, f"{report}: workbook too small"
	# pension uses date/status params
	pen = rep.render_report_xlsx("pension_register", from_date="2016-01-01", to_date="2016-12-31")
	assert pen[:2] == b"PK", "pension: not an xlsx"
	print("rendered            : 7 reports -> valid xlsx")

	# reopen the employee register and check the Arabic header + a totals row
	ws = load_workbook(io.BytesIO(rep.render_report_xlsx("employee_register", run=run.name))).active
	assert ws.cell(row=1, column=1).value == "الموظف", ws.cell(row=1, column=1).value
	assert ws.sheet_view.rightToLeft, "sheet not RTL"
	assert ws.cell(row=ws.max_row, column=1).value == "الإجمالي", "no totals row"
	print("reopened            : header الموظف, RTL, totals row present")

	frappe.db.rollback()                        # discard the run; re-runnable
	print("\nEXCEL EXPORT SMOKE TEST PASSED")


def pdf_export():
	"""Phase 4 M14 — PDF export renders a real (wkhtmltopdf) PDF for every report,
	with bundled Arabic font. Builds a one-employee run, renders each of the 7
	reports to PDF bytes, plus an EMPTY report (no records) still yields a valid
	PDF, then rolls back.
	"""
	from iraqi_government_payroll.api import reports_api as rep

	EMP = "PDF1"
	if not frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": EMP, "employee_name": "PDF Test",
			"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
			"current_stage": 1, "qualification": "Bachelor", "status": "Active",
			"employment_status": "Active", "bank_account": "99001122",
		}).insert()
		frappe.db.commit()
	if not frappe.db.exists("Payroll Period", {"year": 2015, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2015, "month": 6,
						"start_date": "2015-06-01", "end_date": "2015-06-30",
						"status": "Open"}).insert()
		frappe.db.commit()
	period = frappe.get_value("Payroll Period", {"year": 2015, "month": 6}, "name")

	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "Employee",
						  "scope_reference": EMP}).insert()
	run.calculate_run(); run.reload()

	run_reports = ["run_summary", "employee_register", "allowances_register",
				   "deductions_register", "tax_register", "bank_transfer"]
	for report in run_reports:
		pdf = rep.render_report_pdf(report, run=run.name)
		assert pdf[:5] == b"%PDF-", f"{report}: not a PDF"
		assert len(pdf) > 1000, f"{report}: PDF too small ({len(pdf)})"
	pen = rep.render_report_pdf("pension_register", from_date="2015-01-01", to_date="2015-12-31")
	assert pen[:5] == b"%PDF-", "pension: not a PDF"
	print("rendered            : 7 reports -> valid PDF bytes")

	# EMPTY report data still produces a valid PDF (no records in this range)
	empty = rep.render_report_pdf("pension_register", from_date="1990-01-01", to_date="1990-12-31")
	assert empty[:5] == b"%PDF-", "empty: not a PDF"
	assert len(empty) > 1000, "empty PDF too small"
	print("empty report        : valid PDF (%d bytes)" % len(empty))

	frappe.db.rollback()                        # discard the run; re-runnable
	print("\nPDF EXPORT SMOKE TEST PASSED")


def accounting_journal():
	"""Phase 4 M15 — accounting journal export proposes BALANCED rows, fails safely
	on an incomplete mapping, and stays empty/safe for an empty run. Proposal only —
	no GL Entry / Journal Entry is ever created. Rolls back so it is re-runnable.
	"""
	from iraqi_government_payroll.api import accounting_api as acc

	mapping = {
		"salary_expense_account": "5100", "allowance_expense_account": "5200",
		"employee_payable_account": "2100", "pension_payable_account": "2200",
		"tax_payable_account": "2300", "other_deductions_payable_account": "2400",
	}
	for field, value in mapping.items():
		frappe.db.set_single_value("Payroll Account Mapping", field, value)

	EMP = "JRN1"
	if not frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": EMP, "employee_name": "Journal Test",
			"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
			"current_stage": 1, "qualification": "Bachelor", "status": "Active",
			"employment_status": "Active",
		}).insert()
		frappe.db.commit()
	if not frappe.db.exists("Payroll Period", {"year": 2020, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2020, "month": 6,
						"start_date": "2020-06-01", "end_date": "2020-06-30",
						"status": "Open"}).insert()
		frappe.db.commit()
	period = frappe.get_value("Payroll Period", {"year": 2020, "month": 6}, "name")

	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "Employee",
						  "scope_reference": EMP}).insert()
	run.calculate_run(); run.reload()

	j = acc.journal_export(run.name)
	print("journal rows        :", j["rows"])
	print("debit/credit        :", j["total_debit"], "/", j["total_credit"], "| balanced:", j["balanced"])
	assert j["balanced"] is True, j
	assert j["total_debit"] == j["total_credit"], "debit != credit"
	assert j["total_debit"] > 0, "expected a non-trivial journal"
	descs = [r["description"] for r in j["rows"]]
	assert "Salary Expense" in descs and "Employee Payable" in descs, descs
	# debit and credit each sum independently to the same total
	assert sum(r["debit"] for r in j["rows"]) == sum(r["credit"] for r in j["rows"])

	# Incomplete mapping fails safely (no partial/invalid export)
	frappe.db.set_single_value("Payroll Account Mapping", "employee_payable_account", "")
	failed = _blocked(lambda: acc.journal_export(run.name))
	print("incomplete mapping  : fails =", failed)
	assert failed, "incomplete mapping did not fail"
	frappe.db.set_single_value("Payroll Account Mapping", "employee_payable_account", "2100")

	# Empty run -> safe empty balanced journal (no slips for a non-existent employee)
	empty_run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
								"rule_set": "IRAQ-2015", "scope": "Employee",
								"scope_reference": "NOBODY-XYZ"}).insert()
	empty_run.calculate_run(); empty_run.reload()
	ej = acc.journal_export(empty_run.name)
	print("empty journal       :", ej["rows"], "| balanced:", ej["balanced"])
	assert ej["rows"] == [] and ej["balanced"] and ej["total_debit"] == 0, ej

	frappe.db.rollback()                        # discard everything; no posting, re-runnable
	print("\nACCOUNTING JOURNAL SMOKE TEST PASSED")


def security():
	"""Phase 5 M1 — RBAC: roles exist; Payroll Run permissions enforce read-only
	for Read Only User / Auditor and full for the admin; and the sensitive
	accounting-journal export is denied to Read Only User but allowed to Finance.
	"""
	from iraqi_government_payroll.api import accounting_api as acc

	ROLES = ["Government Payroll Administrator", "HR Officer", "Finance Officer", "Read Only User"]
	for r in ROLES:
		assert frappe.db.exists("Role", r), f"missing role: {r}"
	print("roles exist          :", ROLES)

	perms = {p.role: p for p in frappe.get_meta("Payroll Run").permissions}
	ro, ga, fo, au = (perms["Read Only User"], perms["Government Payroll Administrator"],
					  perms["Finance Officer"], perms["Auditor"])
	assert ro.read and not ro.write and not ro.create and not ro.delete and not (ro.export or 0), \
		"Read Only User is not read-only"
	assert ga.read and ga.write and ga.create and ga.delete, "Government Payroll Administrator not full"
	assert fo.read and fo.export and not fo.write, "Finance Officer permissions wrong"
	assert not au.write and not au.create and not au.delete, "Auditor is not read-only"
	print("Payroll Run perms    : ReadOnly read-only · GovAdmin full · Finance read+export · Auditor read-only")

	def ensure_user(email, roles):
		if not frappe.db.exists("User", email):
			frappe.get_doc({"doctype": "User", "email": email, "send_welcome_email": 0,
							"first_name": email.split("@")[0]}).insert(ignore_permissions=True)
		u = frappe.get_doc("User", email)
		have = {r.role for r in u.roles}
		for r in roles:
			if r not in have:
				u.append("roles", {"role": r})
		u.save(ignore_permissions=True)
		return email

	ro_user = ensure_user("sec_readonly@test.local", ["Read Only User"])
	fin_user = ensure_user("sec_finance@test.local", ["Finance Officer"])
	frappe.db.commit()

	# Live permission layer
	assert frappe.has_permission("Payroll Run", "read", user=ro_user), "Read Only User cannot read"
	assert not frappe.has_permission("Payroll Run", "write", user=ro_user), "Read Only User can WRITE"
	print("Read Only User (live): read=yes write=no")

	# Sensitive-action gate via the real endpoint (set_user)
	frappe.set_user(ro_user)
	ro_denied = _blocked(lambda: acc.journal_export("NO-SUCH-RUN"))
	frappe.set_user("Administrator")
	assert ro_denied, "Read Only User was ALLOWED to export the accounting journal"
	print("journal export gate  : Read Only User denied")

	frappe.set_user(fin_user)
	fin_result = acc.journal_export("NO-SUCH-RUN")     # gate passes; empty run -> balanced
	frappe.set_user("Administrator")
	assert fin_result["balanced"], "Finance Officer export failed"
	print("journal export gate  : Finance Officer allowed (empty balanced)")

	frappe.db.rollback()
	print("\nSECURITY SMOKE TEST PASSED")


def grade_validation():
	"""Phase 5 M4.1 — Government Grade master + scale-placement validation, live.

	Verifies, end to end on the bench:
	  1. the Government Grade master holds the 13 legacy codes (values preserved);
	  2. a profile saved with only grade_code backfills the new grade Link (sync);
	  3. an invalid placement (grade 7 stage 99) is REJECTED at profile save — i.e.
	     before any payroll calculation;
	  4. a promotion changes the employee grade Link (and grade_code mirror) WITHOUT
	     touching Government Position / Current Position (salary source = the scale,
	     not the Position);
	  5. an annual increment changes only the employee-level stage, never the grade.
	Rolls back at the end so it is re-runnable.
	"""
	from iraqi_government_payroll.services.payroll_engine import repository as repo

	# 1. Master preserves the 13 legacy codes exactly.
	master = set(frappe.get_all("Government Grade", pluck="name"))
	expected = {"10", "9", "8", "7", "6", "5", "4", "3", "2", "1",
				"SPECIAL_A", "SPECIAL_B", "SPECIAL_C"}
	print("grade master         :", sorted(master))
	assert master >= expected, f"missing grade codes: {expected - master}"

	# 2. Saving with only grade_code backfills the grade Link (controller sync).
	EMP = "GRD1"
	if frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.delete_doc("Government Employee Payroll Profile", EMP, force=True)
	prof = frappe.get_doc({
		"doctype": "Government Employee Payroll Profile",
		"employee_number": EMP, "employee_name": "Grade Test",
		"rule_set": "IRAQ-2015", "grade_code": "7", "current_grade": 7,
		"current_stage": 1, "qualification": "Bachelor", "status": "Active",
		"employment_status": "Active",
		"appointment_date": "2000-01-01",
		"current_grade_date": "2000-01-01", "current_stage_date": "2000-01-01",
	}).insert()
	prof.reload()
	print("grade backfilled     : grade=%r grade_code=%r" % (prof.grade, prof.grade_code))
	assert prof.grade == "7", f"grade Link not synced from grade_code: {prof.grade!r}"

	# 3. Invalid placement is rejected at SAVE (before any payroll run).
	prof.current_stage = 99
	rejected = _blocked(lambda: prof.save())
	print("invalid stage 99     : rejected =", rejected)
	assert rejected, "grade 7 stage 99 was accepted — validation did not fire"
	prof.reload()                                # discard the bad in-memory change
	assert prof.current_stage == 1, "profile mutated despite rejected save"

	# Capture Position before promotion (must be untouched by the promotion).
	pos_before = (prof.government_position, prof.current_position)

	# 4. Promotion changes grade WITHOUT Position. Grade 7 -> 6 per the scale.
	preq = frappe.get_doc({
		"doctype": "Promotion Request", "employee_profile": EMP,
		"employee_name": "Grade Test", "approval_status": "Approved",
		"due_date": "2024-01-01",
	}).insert()
	res = repo.apply_promotion(preq)        # mutates preq in memory (caller persists)
	prof.reload()
	print("after promotion      : grade=%r grade_code=%r stage=%r (was 7)" %
		  (prof.grade, prof.grade_code, prof.current_stage))
	assert prof.grade != "7", "promotion did not change the grade"
	assert prof.grade == prof.grade_code, "grade Link and grade_code mirror diverged"
	assert frappe.db.exists("Government Grade", prof.grade), "new grade not in master"
	assert (prof.government_position, prof.current_position) == pos_before, \
		"promotion changed Government Position — grade must not derive from Position"
	print("position unchanged   :", pos_before, "(grade driven by scale, not Position)")
	# M4.2: the request records grade transitions as Links to the master
	print("promotion grade Links: from_grade_ref=%r to_grade_ref=%r" %
		  (preq.from_grade_ref, preq.to_grade_ref))
	assert preq.from_grade_ref == "7" and preq.to_grade_ref == prof.grade, \
		"promotion from/to_grade_ref Links not set to master grades"

	# 5. Increment changes only the employee-level stage, not the grade.
	grade_after_promo = prof.grade
	stage_before = prof.current_stage
	ireq = frappe.get_doc({
		"doctype": "Annual Increment Request", "employee_profile": EMP,
		"employee_name": "Grade Test", "approval_status": "Approved",
		"due_date": "2025-01-01",
	}).insert()
	repo.apply_increment(ireq)              # mutates ireq in memory (caller persists)
	prof.reload()
	print("after increment      : grade=%r stage=%r (was %r) | req.current_grade_ref=%r" %
		  (prof.grade, prof.current_stage, stage_before, ireq.current_grade_ref))
	assert prof.grade == grade_after_promo, "increment changed the grade — it must not"
	assert prof.current_stage == stage_before + 1, "increment did not advance the stage by 1"
	assert ireq.current_grade_ref == grade_after_promo, "increment current_grade_ref Link wrong"

	# 6. Government Position owns NO grade or stage field (organizational only).
	pos_fields = {f.fieldname for f in frappe.get_meta("Government Position").fields}
	grade_stage = {f for f in pos_fields if "grade" in f.lower() or "stage" in f.lower()}
	print("position grade/stage : %s | has position_type: %s" %
		  (grade_stage or "none", "position_type" in pos_fields))
	assert not grade_stage, f"Government Position must not own grade/stage: {grade_stage}"
	assert "position_type" in pos_fields, "position_type missing on Government Position"

	frappe.db.rollback()                         # discard everything; re-runnable
	print("\nGRADE VALIDATION SMOKE TEST PASSED")


def uat_demo_cycle():
	"""UAT Demo Cycle — exercise the system as a real payroll user across personas.

	Creates clearly-marked UAT-* demo employees (committed, idempotent), then runs
	the full cycle: profiles, salary (single/married/family), invalid-profile
	rejection, increment/promotion eligibility (apply), pension service-year sweep,
	a Draft->Review->Approve->Lock payroll run with totals from the snapshot and
	lock immutability, and locked-run reports from the snapshot. The transactional
	steps roll back, so the personas persist but nothing else does (re-runnable).
	"""
	from iraqi_government_payroll.services.payroll_engine import repository as repo
	from iraqi_government_payroll.services.pension.pension_service import (
		compute_retirement_pension, RetirementPensionInput)
	from iraqi_government_payroll.api import reports_api as rep

	today = frappe.utils.nowdate()
	yago = lambda n: frappe.utils.add_years(today, -n)
	ENT = "UAT-ENTITY"

	# ---- 1. Personas (committed, idempotent) ----
	if not frappe.db.exists("Government Entity", ENT):
		frappe.get_doc({"doctype": "Government Entity", "entity_code": ENT,
						"entity_name_ar": "دائرة الاختبار (UAT)", "entity_type": "Ministry"}
					   ).insert(ignore_permissions=True)

	def persona(num, name, grade, stage, qual, marital, *, stage_date=None,
				grade_date=None, appoint=None, children=0):
		if frappe.db.exists("Government Employee Payroll Profile", num):
			return
		doc = {"doctype": "Government Employee Payroll Profile", "employee_number": num,
			   "employee_name": name, "rule_set": "IRAQ-2015", "grade": grade,
			   "current_stage": stage, "qualification": qual, "status": "Active",
			   "employment_status": "Active", "government_entity": ENT,
			   "marital_status": marital, "appointment_date": appoint or yago(5),
			   "current_stage_date": stage_date or yago(1), "current_grade_date": grade_date or yago(2),
			   "national_id": "199" + num.replace("UAT-", "").replace("-", "")[:9].ljust(9, "0")}
		if children:
			doc["family_members"] = [
				{"full_name": f"ابن {i}", "relation": "Son", "date_of_birth": yago(8 + i),
				 "is_alive": 1, "financially_dependent": 1} for i in range(children)]
		frappe.get_doc(doc).insert()

	persona("UAT-01-SINGLE", "أعزب (UAT)", "7", 1, "Bachelor", "Single")
	persona("UAT-02-MARRIED", "متزوج (UAT)", "7", 1, "Bachelor", "Married")
	persona("UAT-03-FAMILY", "متزوج بأطفال (UAT)", "7", 1, "Bachelor", "Married", children=3)
	persona("UAT-04-INCR-Y", "علاوة مستحقة (UAT)", "7", 2, "Bachelor", "Single", stage_date=yago(2))
	persona("UAT-05-INCR-N", "علاوة غير مستحقة (UAT)", "7", 3, "Bachelor", "Single", stage_date=today)
	persona("UAT-06-PROMO-Y", "ترفيع مستحق (UAT)", "7", 1, "Bachelor", "Single", grade_date=yago(5))
	persona("UAT-07-PROMO-N", "ترفيع غير مستحق (UAT)", "7", 1, "Bachelor", "Single", grade_date=today)
	persona("UAT-08-RETIRE", "قرب التقاعد (UAT)", "3", 5, "Master", "Married", appoint=yago(38))
	frappe.db.commit()
	print("personas             : 8 UAT employees + entity ready")

	# ---- 2. Profile UAT ----
	p3 = frappe.get_doc("Government Employee Payroll Profile", "UAT-03-FAMILY")
	assert p3.government_entity == ENT and p3.grade == "7" and p3.current_stage == 1
	assert p3.marital_status == "Married" and p3.eligible_children_count == 3, p3.eligible_children_count
	print("profile UAT          : entity/grade/stage/marital/children verified (family=%d)" % p3.eligible_children_count)

	# ---- 3. Salary preview + invalid-profile rejection (Arabic) ----
	from iraqi_government_payroll.api import payroll_api as papi
	prev = papi.salary_preview("IRAQ-2015", "7", 1)
	assert prev["valid"] and prev["basic_salary"] == 296000
	bad = papi.salary_preview("IRAQ-2015", "7", 99)
	assert not bad["valid"] and "سلم الرواتب" in bad["message"]
	def _bad_profile():
		frappe.get_doc({"doctype": "Government Employee Payroll Profile",
						"employee_number": "UAT-BAD", "employee_name": "x", "rule_set": "IRAQ-2015",
						"grade": "7", "current_stage": 99, "status": "Active"}).insert()
	assert _blocked(_bad_profile), "invalid profile was accepted"
	print("salary/invalid       : preview 296000; invalid stage 99 rejected (Arabic)")

	# ---- 4. Increment eligibility (apply via engine; rolled back) ----
	def _apply_increment(emp):
		req = frappe.get_doc({"doctype": "Annual Increment Request", "employee_profile": emp,
							  "approval_status": "Draft"}).insert()
		repo.apply_increment(req)
		return frappe.db.get_value("Government Employee Payroll Profile", emp, "current_stage")
	new_stage = _apply_increment("UAT-04-INCR-Y")
	print("increment eligible   : UAT-04 stage 2 ->", new_stage)
	assert new_stage == 3, f"eligible increment did not advance the stage ({new_stage})"
	assert _blocked(lambda: _apply_increment("UAT-05-INCR-N")), "not-eligible increment applied"
	frappe.db.rollback()

	# ---- 5. Promotion eligibility (apply via engine; rolled back) ----
	def _apply_promotion(emp):
		req = frappe.get_doc({"doctype": "Promotion Request", "employee_profile": emp,
							  "approval_status": "Draft"}).insert()
		repo.apply_promotion(req)
		return frappe.db.get_value("Government Employee Payroll Profile", emp, "grade")
	new_grade = _apply_promotion("UAT-06-PROMO-Y")
	print("promotion eligible   : UAT-06 grade 7 ->", new_grade)
	assert new_grade == "6", f"eligible promotion did not change the grade ({new_grade})"
	assert _blocked(lambda: _apply_promotion("UAT-07-PROMO-N")), "not-eligible promotion applied"
	frappe.db.rollback()

	# ---- 6. Pension service-year sweep (pure engine) ----
	pr, certs, brackets = repo.load_pension_rule_data("IRAQ-2015")
	def _pension(years):
		return compute_retirement_pension(RetirementPensionInput(
			avg36=1000000, service_years=years, last_functional_salary=1000000,
			last_full_salary=1000000, qualification="Master",
			cost_of_living_method="Fixed Percentage", cost_of_living_value=30), pr, certs, brackets)
	pens = {y: _pension(y) for y in (15, 20, 30, 35)}
	print("pension sweep        : approved", {y: pens[y].approved_pension for y in pens})
	assert [pens[y].approved_pension for y in (15, 20, 30, 35)] == sorted(pens[y].approved_pension for y in (15, 20, 30, 35))
	assert pens[15].end_of_service_bonus == 0 and pens[30].end_of_service_bonus == 12000000
	assert all(pens[y].certificate_allowance > 0 and pens[y].cost_of_living > 0 and pens[y].net_pension > 0 for y in pens)

	# ---- 7. Payroll Run: Draft -> Review -> Approve -> Lock (rolled back) ----
	if not frappe.db.exists("Payroll Period", {"year": 2027, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2027, "month": 6,
						"start_date": "2027-06-01", "end_date": "2027-06-30", "status": "Open"}).insert()
		frappe.db.commit()
	period = frappe.get_value("Payroll Period", {"year": 2027, "month": 6}, "name")
	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period, "rule_set": "IRAQ-2015",
						  "scope": "Government Entity", "scope_reference": ENT}).insert()
	run.calculate_run(); run.reload()
	assert run.workflow_state == "Calculated" and run.total_employees == 8, (run.workflow_state, run.total_employees)
	# salary correctness across personas (single < married < family)
	def _net(emp):
		return frappe.db.get_value("Salary Slip", {"payroll_run": run.name, "employee_profile": emp}, "net_salary")
	n1, n2, n3 = _net("UAT-01-SINGLE"), _net("UAT-02-MARRIED"), _net("UAT-03-FAMILY")
	print("salary run           : single=%s married=%s family=%s" % (n1, n2, n3))
	assert n1 < n2 < n3, "family allowances did not raise net (single<married<family)"
	assert frappe.db.get_value("Salary Slip", {"payroll_run": run.name, "employee_profile": "UAT-01-SINGLE"}, "basic_salary") == 296000
	# submit slips (write immutable snapshots) then drive governance to Locked
	for sl in frappe.get_all("Salary Slip", filters={"payroll_run": run.name, "docstatus": 0}, pluck="name"):
		frappe.get_doc("Salary Slip", sl).submit()
	run.submit_for_review(); run.approve_run(); run.submit_run(); run.lock_run(); run.reload()
	assert run.workflow_state == "Locked"
	# totals from the immutable snapshot; lock prevents recalculation
	summary = rep.run_summary(run.name)
	assert summary["employees"] == 8 and summary["total_net"] > 0
	assert _blocked(lambda: run.calculate_run()), "locked run was recalculated"
	print("payroll run          : 8 employees, locked, total_net=%s, recalc blocked" % summary["total_net"])

	# ---- 8. Reports from the immutable snapshot (locked run) ----
	er = rep.employee_register(run.name)
	bt = rep.bank_transfer(run.name)
	from iraqi_government_payroll.api import accounting_api as acc
	jr = acc.journal_export(run.name)
	assert len(er["rows"]) == 8 and len(bt["rows"]) == 8 and jr["balanced"]
	assert er["totals"]["net"] == summary["total_net"], "report net != summary net (snapshot)"
	print("reports (locked)     : register/bank 8 rows from snapshot, journal balanced")

	frappe.db.rollback()                      # discard run/slips/snapshots; personas persist
	print("\nUAT DEMO CYCLE SMOKE TEST PASSED")


def transaction_forms():
	"""Phase 5 M8 — the frontend transaction-create endpoints: increment/promotion
	drafts and a pension calculation (computed by the existing engine). Cleans up
	the created records (re-runnable)."""
	from iraqi_government_payroll.api import payroll_api as papi

	emp = frappe.db.get_value("Government Employee Payroll Profile",
							  {"employee_number": ["like", "DEMO-EMP-%"]}, "name") \
		or frappe.db.get_value("Government Employee Payroll Profile", {}, "name")
	assert emp, "no employee profile to test with"

	inc = papi.create_increment_request(emp, due_date="2025-01-01", remarks="smoke")
	print("increment draft      :", inc)
	assert inc["approval_status"] == "Draft" and frappe.db.exists("Annual Increment Request", inc["name"])

	promo = papi.create_promotion_request(emp, vacancy_available=1, direct_manager_recommendation=1,
										  committee_decision="موافقة", remarks="smoke")
	print("promotion draft      :", promo)
	assert promo["approval_status"] == "Draft" and frappe.db.exists("Promotion Request", promo["name"])

	pen = papi.create_pension_calculation(emp, service_years=30, average_36_months=1000000,
										  last_functional_salary=1000000, last_full_salary=1000000)
	print("pension calc         : net=%s gross=%s approved=%s eos=%s" %
		  (pen["net_pension"], pen["gross_pension"], pen["approved_pension"], pen["end_of_service_bonus"]))
	assert pen["net_pension"] > 0 and pen["gross_pension"] >= pen["approved_pension"]
	assert pen["end_of_service_bonus"] == 1000000 * 12, "EOS at 30 years = last salary x 12"
	pdoc = frappe.get_doc("Pension Calculation", pen["name"])
	assert pdoc.net_pension == pen["net_pension"], "stored net != computed net"

	# cleanup (drafts are deletable)
	frappe.delete_doc("Annual Increment Request", inc["name"], force=True)
	frappe.delete_doc("Promotion Request", promo["name"], force=True)
	frappe.delete_doc("Pension Calculation", pen["name"], force=True)
	frappe.db.commit()
	print("\nTRANSACTION FORMS SMOKE TEST PASSED")


def data_integrity():
	"""Backend functional — Data Integrity (item 1): active rule set, grade master,
	salary-scale coverage, grade/stage validation, profile required fields, and
	position/entity linkage. Read-only + rolled back (re-runnable)."""
	# Active rule set
	active_rs = frappe.get_all("Government Rule Set", filters={"status": "Active"}, pluck="name")
	print("active rule sets     :", active_rs)
	assert "IRAQ-2015" in active_rs, "no active IRAQ-2015 rule set"

	# Government grades master
	grades = frappe.get_all("Government Grade", filters={"active": 1}, pluck="name")
	assert len(grades) == 13, f"expected 13 active grades, got {len(grades)}"

	# Salary scale coverage: an active scale exists and every grade/stage has a basic
	scale = frappe.db.get_value("Government Salary Scale", {"rule_set": "IRAQ-2015", "is_active": 1}, "name") \
		or frappe.db.get_value("Government Salary Scale", {"rule_set": "IRAQ-2015"}, "name")
	assert scale, "no salary scale for IRAQ-2015"
	details = frappe.get_all("Government Salary Scale Detail", filters={"parent": scale},
							 fields=["grade_code", "stage", "basic_salary"])
	assert len(details) >= 130, f"thin scale coverage: {len(details)} rows"
	assert all(d.basic_salary and d.basic_salary > 0 for d in details), "scale row with no basic salary"
	print("scale coverage       :", len(details), "grade/stage rows, all priced")

	# Grade/stage validation: an invalid placement is rejected with an Arabic message
	EMP = "DI1"
	if frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.delete_doc("Government Employee Payroll Profile", EMP, force=True)
	def _bad_stage():
		frappe.get_doc({"doctype": "Government Employee Payroll Profile",
						"employee_number": EMP, "employee_name": "DI", "rule_set": "IRAQ-2015",
						"grade": "7", "current_stage": 99, "status": "Active"}).insert()
	assert _blocked(_bad_stage), "invalid (grade, stage) was accepted"

	# Profile required fields: grade is mandatory
	def _missing_grade():
		frappe.get_doc({"doctype": "Government Employee Payroll Profile",
						"employee_number": EMP, "employee_name": "DI",
						"rule_set": "IRAQ-2015", "current_stage": 1, "status": "Active"}).insert()
	assert _blocked(_missing_grade), "profile without a grade was accepted"
	print("validation           : invalid stage + missing grade both rejected")

	# Position + Entity linkage: link both on a profile and resolve their names.
	from iraqi_government_payroll.api.slip_api import _position_title
	if not frappe.db.exists("Government Position", "POS-DI"):
		frappe.get_doc({"doctype": "Government Position", "position_code": "POS-DI",
						"position_name_ar": "مدير القسم (تجريبي)", "position_type": "Director"}
					   ).insert(ignore_permissions=True)
	ent = frappe.db.get_value("Government Entity", {}, "name")
	assert ent, "no Government Entity exists"
	prof = frappe.get_doc({"doctype": "Government Employee Payroll Profile",
						   "employee_number": EMP, "employee_name": "DI", "rule_set": "IRAQ-2015",
						   "grade": "7", "current_stage": 1, "status": "Active",
						   "government_position": "POS-DI", "government_entity": ent}).insert()
	title = _position_title(prof.as_dict())
	ent_ar = frappe.db.get_value("Government Entity", ent, "entity_name_ar")
	print("linkage              : position->%s | entity %s->%s" % (title, ent, ent_ar))
	assert title == "مدير القسم (تجريبي)", f"position link not resolved: {title}"
	assert ent_ar, "Government Entity has no Arabic name"

	frappe.db.rollback()
	print("\nDATA INTEGRITY SMOKE TEST PASSED")


def financial_wiring():
	"""Backend financial-wiring check: a real payroll calculation reconciles
	net = earnings − deductions from the slip, and the pre-calc guard blocks a run
	with no eligible employees. Rolls back the runs (re-runnable)."""
	EMP = "FW1"
	if not frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": EMP, "employee_name": "Financial Wiring Test",
			"rule_set": "IRAQ-2015", "grade": "7", "current_stage": 1,
			"qualification": "Bachelor", "status": "Active", "employment_status": "Active",
		}).insert()
		frappe.db.commit()
	if not frappe.db.exists("Payroll Period", {"year": 2021, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2021, "month": 6,
						"start_date": "2021-06-01", "end_date": "2021-06-30", "status": "Open"}).insert()
		frappe.db.commit()
	period = frappe.get_value("Payroll Period", {"year": 2021, "month": 6}, "name")

	run = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "IRAQ-2015", "scope": "Employee", "scope_reference": EMP}).insert()
	run.calculate_run(); run.reload()
	assert run.workflow_state == "Calculated", run.workflow_state
	slip = frappe.get_doc("Salary Slip", frappe.db.get_value(
		"Salary Slip", {"payroll_run": run.name, "employee_profile": EMP}, "name"))
	print("basic/earnings/deductions/net:", slip.basic_salary, slip.total_earnings,
		  slip.total_deductions, slip.net_salary)
	assert slip.basic_salary == 296000, slip.basic_salary              # g7s1 Bachelor anchor
	assert slip.total_earnings == slip.basic_salary + slip.total_capped_allowances + slip.total_non_capped_allowances
	assert slip.net_salary == slip.total_earnings - slip.total_deductions, "net != earnings - deductions"
	assert slip.net_salary > 0

	# Pre-calc guard (item 11): a rule set with no active salary scale is blocked
	# with a clear message — not per-employee failures.
	if not frappe.db.exists("Government Rule Set", "FW-NOSCALE"):
		frappe.get_doc({"doctype": "Government Rule Set", "rule_set_code": "FW-NOSCALE",
						"rule_set_name": "No Scale (test)", "year": 2099,
						"effective_from": "2099-01-01", "status": "Draft"}).insert(ignore_permissions=True)
		frappe.db.commit()
	bad = frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
						  "rule_set": "FW-NOSCALE", "scope": "All"}).insert()
	blocked = _blocked(lambda: bad.calculate_run())
	print("no-scale guard       : blocked =", blocked)
	assert blocked, "pre-calc guard did not block a rule set with no active salary scale"

	frappe.db.rollback()                          # discard the runs/slips; re-runnable
	print("\nFINANCIAL WIRING SMOKE TEST PASSED")


def family_dependents():
	"""Phase 5 M7 — Family & Dependents: the API saves dependents, the controller
	recomputes ages/eligibility/counts, the engine reads eligible_children_count to
	scale the configurable Family/Child allowance, and the family state is recorded
	into the snapshot input. Re-runnable (deletes its employee).
	"""
	from iraqi_government_payroll.api import payroll_api as papi
	from iraqi_government_payroll.services.payroll_engine import repository as repo
	from iraqi_government_payroll.services.audit.audit_service import build_snapshot_payload

	EMP = "FAM1"
	if frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.delete_doc("Government Employee Payroll Profile", EMP, force=True)
		frappe.db.commit()

	today = frappe.utils.nowdate()
	def yago(n):
		return frappe.utils.add_years(today, -n)

	papi.save_employee_profile({
		"employee_number": EMP, "employee_name": "Family Test",
		"rule_set": "IRAQ-2015", "grade": "7", "current_stage": 1,
		"qualification": "Bachelor", "status": "Active", "marital_status": "Married",
		"family_members": [
			{"full_name": "Zawja", "relation": "Spouse", "date_of_birth": yago(35)},
			{"full_name": "Child A", "relation": "Son", "date_of_birth": yago(10)},
			{"full_name": "Child B", "relation": "Son", "date_of_birth": yago(25)},
			{"full_name": "Child C", "relation": "Daughter", "date_of_birth": yago(20), "is_student": 1},
			{"full_name": "Child D", "relation": "Son", "date_of_birth": yago(30),
			 "is_employed": 1, "employment_type": "Government", "monthly_income": 500000},
		],
	})
	p = frappe.get_doc("Government Employee Payroll Profile", EMP)
	print("counts               : spouse=%d children=%d eligible_children=%d employed=%d student=%d eligible_dep=%d" %
		  (p.spouse_count, p.children_count, p.eligible_children_count,
		   p.employed_dependents_count, p.student_dependents_count, p.eligible_dependents_count))
	assert p.spouse_count == 1, p.spouse_count
	assert p.children_count == 4, p.children_count          # A,B,C,D
	assert p.eligible_children_count == 2, p.eligible_children_count  # A(10) + C(20 student); B(25), D(employed) excluded
	assert p.employed_dependents_count == 1, p.employed_dependents_count
	assert p.student_dependents_count == 1, p.student_dependents_count
	assert p.eligible_dependents_count == 3, p.eligible_dependents_count  # spouse + A + C

	# ages + per-member eligibility written back
	by = {m.full_name: m for m in p.family_members}
	assert by["Child A"].age == 10 and by["Child A"].eligible_for_family_allowance == 1
	assert by["Child D"].eligible_for_family_allowance == 0, "employed dependent must be ineligible"

	# engine reads eligible_children_count -> Family/Child allowance scales with it
	res = repo.calculate_for_profile(EMP, "2020-06-01")
	lines = {l.component_code: l.amount for l in res.allowance_lines}
	fam_child_unit = frappe.db.get_value("Allowance Rule", "FAM_CHILD", "fixed_amount")
	print("family allowances    : FAM_CHILD=%s FAM_SPOUSE=%s (unit=%s)" %
		  (lines.get("FAM_CHILD"), lines.get("FAM_SPOUSE"), fam_child_unit))
	assert lines.get("FAM_CHILD") == fam_child_unit * min(p.eligible_children_count, 4)
	assert lines.get("FAM_SPOUSE"), "married -> spouse allowance expected"

	# family state recorded into the snapshot input (immutable history)
	emp_input = repo._employee_input_from_profile(p.as_dict(), "2020-06-01")
	snap = build_snapshot_payload(res, employee_input=emp_input, employee_profile=EMP)
	import json as _json
	fs = _json.loads(snap["input_snapshot"]).get("family_summary") or {}
	print("snapshot family_summary:", fs)
	assert fs.get("eligible_children_count") == 2, fs

	frappe.delete_doc("Government Employee Payroll Profile", EMP, force=True)
	frappe.db.commit()
	print("\nFAMILY DEPENDENTS SMOKE TEST PASSED")


def async_payroll_run():
	"""Production blocker H1 — asynchronous (queued) payroll calculation.

	Proves, without changing any calculation logic:
	  1. the synchronous calculate_run still works (small run);
	  2. calculate_run_async marks the run Queued and enqueues a background job;
	  3. the queued worker path (jobs.run_calculation_job) produces an IDENTICAL
	     immutable result (same engine path, same net, error-isolation tally intact);
	  4. a locked run cannot be recalculated via the async entry point.

	Per-employee failure isolation itself is unchanged (the job reuses
	repository.run_payroll -> run_payroll_batch) and is covered by the host test
	test_payroll_run.test_mixed_error_completed_with_warnings.
	Cleans up its runs/slips (re-runnable).
	"""
	from iraqi_government_payroll.services.payroll_engine import jobs, governance

	EMP = "ASYNC1"
	if not frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.get_doc({
			"doctype": "Government Employee Payroll Profile",
			"employee_number": EMP, "employee_name": "Async Run Test",
			"rule_set": "IRAQ-2015", "grade": "7", "current_stage": 1,
			"qualification": "Bachelor", "status": "Active", "employment_status": "Active",
		}).insert()
		frappe.db.commit()
	if not frappe.db.exists("Payroll Period", {"year": 2028, "month": 6}):
		frappe.get_doc({"doctype": "Payroll Period", "year": 2028, "month": 6,
						"start_date": "2028-06-01", "end_date": "2028-06-30", "status": "Open"}).insert()
		frappe.db.commit()
	period = frappe.get_value("Payroll Period", {"year": 2028, "month": 6}, "name")

	def _new_run():
		return frappe.get_doc({"doctype": "Payroll Run", "payroll_period": period,
							   "rule_set": "IRAQ-2015", "scope": "Employee",
							   "scope_reference": EMP}).insert()

	created_runs = []

	# 1) Synchronous path still works (small run).
	sync_run = _new_run(); created_runs.append(sync_run.name)
	sync_run.calculate_run(); sync_run.reload()
	assert sync_run.workflow_state == "Calculated", sync_run.workflow_state
	sync_net = frappe.db.get_value(
		"Salary Slip", {"payroll_run": sync_run.name, "employee_profile": EMP}, "net_salary")
	print("sync run            :", sync_run.workflow_state, "| net:", sync_net)
	assert sync_net and sync_net > 0, sync_net

	# 2) Async path: enqueue -> Queued + a background job is scheduled.
	async_run = _new_run(); created_runs.append(async_run.name)
	res = async_run.calculate_run_async()
	async_run.reload()
	print("enqueue result      :", res, "| run_status:", async_run.run_status)
	assert res.get("status") == "queued", res
	assert async_run.run_status == "Queued", async_run.run_status
	assert async_run.workflow_state == "Draft", async_run.workflow_state  # not yet calculated
	# Don't let the deferred (after-commit) enqueue hit RQ in this test process.
	frappe.flags.enqueue_after_commit = []

	# 3) Run the worker entry point directly -> identical immutable result.
	jobs.run_calculation_job(async_run.name, frappe.session.user)
	async_run.reload()
	async_net = frappe.db.get_value(
		"Salary Slip", {"payroll_run": async_run.name, "employee_profile": EMP}, "net_salary")
	print("worker result       :", async_run.workflow_state, "| run_status:",
		  async_run.run_status, "| net:", async_net,
		  "| success/err:", async_run.success_count, async_run.error_count)
	assert async_run.workflow_state == "Calculated", async_run.workflow_state
	assert async_run.run_status in ("Completed", "Completed With Warnings"), async_run.run_status
	assert async_net == sync_net, (async_net, sync_net)       # queued == sync, no logic drift
	# Isolation tally intact: the one employee was processed with no blocking error
	# (a non-blocking warning is fine — the run is still produced).
	assert (async_run.error_count or 0) == 0, async_run.error_count
	assert (async_run.processed_count or 0) == 1, async_run.processed_count
	assert ((async_run.success_count or 0) + (async_run.warning_count or 0)) == 1, \
		(async_run.success_count, async_run.warning_count)

	# 4) A locked run cannot be recalculated via the async entry point (same
	#    governance.ensure_can_calculate guard as the sync path).
	locked_guard = _new_run()
	locked_guard.workflow_state = governance.LOCKED            # in-memory; the guard reads it
	blocked = _blocked(lambda: locked_guard.calculate_run_async())
	print("locked async block  :", blocked)
	assert blocked, "calculate_run_async must refuse a locked run"
	assert locked_guard.run_status != "Queued", "locked run must not be marked Queued"

	# Cleanup: delete draft slips + runs (no snapshot yet — slips are draft).
	for rn in created_runs:
		for sl in frappe.get_all("Salary Slip", filters={"payroll_run": rn}, pluck="name"):
			frappe.delete_doc("Salary Slip", sl, force=True)
		if frappe.db.exists("Payroll Run", rn):
			frappe.delete_doc("Payroll Run", rn, force=True)
	frappe.db.commit()
	print("\nASYNC PAYROLL RUN SMOKE TEST PASSED")


def payroll_slip():
	"""Phase 5 M6 — generate the Government Payroll Slip from a post-calc snapshot
	(source of truth) and render its RTL Arabic PDF. Idempotent; cleans up the
	generated slip so it is re-runnable. No payroll recalculation.
	"""
	from iraqi_government_payroll.api import slip_api
	from iraqi_government_payroll.services.reports.slip_pdf import build_slip_html, render_slip_pdf

	# A snapshot of a calculated Salary Slip = the source of truth for the amounts.
	snap = frappe.db.get_value(
		"Payroll Calculation Snapshot", {"calculation_type": "Salary Slip"},
		["name", "salary_slip", "net_amount"], as_dict=True)
	assert snap and snap.salary_slip, "no calculated Salary Slip snapshot found"
	slip = frappe.get_doc("Salary Slip", snap.salary_slip)

	res = slip_api.generate_payroll_slip(snap.salary_slip)
	print("generated            :", res)
	assert res["source"] == "snapshot", res
	doc = frappe.get_doc("Government Payroll Slip", res["name"])

	# Net comes straight from the snapshot/slip — never recomputed here.
	print("net (slip/snapshot)  :", doc.net_pay, "/", slip.net_salary)
	assert doc.net_pay == slip.net_salary, (doc.net_pay, slip.net_salary)
	assert doc.snapshot == snap.name
	# Entitlement = base + allowances (+ print-only adjustment/rewards, default 0)
	calc_ent = doc.base_salary + doc.total_allowances + (doc.adjustment_amount or 0) + (doc.total_rewards or 0)
	assert doc.total_entitlement == calc_ent, (doc.total_entitlement, calc_ent)
	assert doc.amount_before_rounding == doc.net_pay
	assert (doc.amount_after_rounding or 0) % 250 == 0, doc.amount_after_rounding
	# allowance lines exclude the basic salary; deduction lines come from the snapshot
	assert all(a.amount for a in doc.allowance_lines), "empty allowance line"
	print("lines                : allowances=%d deductions=%d misc=%d" %
		  (len(doc.allowance_lines), len(doc.deduction_lines), len(doc.misc_deduction_lines)))
	print("identity             : %s | grade %s/%s | %s years service" %
		  (doc.employee_number, doc.grade, doc.stage, doc.years_of_service))

	# Printable RTL PDF
	pdf = render_slip_pdf(build_slip_html(doc.as_dict()))
	print("pdf                  :", pdf[:5], len(pdf), "bytes")
	assert pdf[:5] == b"%PDF-" and len(pdf) > 1000, "invalid slip PDF"

	# Idempotent regenerate (upsert, one slip per Salary Slip)
	res2 = slip_api.generate_payroll_slip(snap.salary_slip)
	assert res2["name"] == res["name"], "generation is not idempotent"

	frappe.delete_doc("Government Payroll Slip", res["name"], force=True)
	frappe.db.commit()
	print("\nPAYROLL SLIP SMOKE TEST PASSED")


def employee_profile_api():
	"""Phase 5 M5 — the frontend employee create/edit endpoints: master list,
	salary preview, create/update, RBAC, and Arabic validation. Uses a dedicated
	employee number and cleans up at the end so it is re-runnable.
	"""
	from iraqi_government_payroll.api import payroll_api as papi

	# list_grades -> the 13 active master codes
	grades = papi.list_grades()
	codes = {g["grade_code"] for g in grades}
	print("active grades        :", sorted(codes))
	assert codes >= {"7", "6", "SPECIAL_A"} and len(grades) == 13, grades
	assert all(g.get("grade_name_ar") for g in grades), "grade missing Arabic name"

	# salary_preview: valid combo returns a basic; invalid returns Arabic message
	ok = papi.salary_preview("IRAQ-2015", "7", 1)
	print("preview 7/1          :", ok)
	assert ok["valid"] and ok["basic_salary"] and ok["basic_salary"] > 0, ok
	bad = papi.salary_preview("IRAQ-2015", "7", 99)
	print("preview 7/99         :", bad["valid"], "|", bad["message"])
	assert not bad["valid"] and bad["basic_salary"] is None, bad
	assert "سلم الرواتب" in bad["message"], "expected an Arabic scale message"

	EMP = "GRD-API-1"
	if frappe.db.exists("Government Employee Payroll Profile", EMP):
		frappe.delete_doc("Government Employee Payroll Profile", EMP, force=True)
		frappe.db.commit()

	# create via the endpoint (Administrator) — grade Link only, no deprecated fields
	created = papi.save_employee_profile({
		"employee_number": EMP, "employee_name": "API Create Test",
		"rule_set": "IRAQ-2015", "grade": "7", "current_stage": 1,
		"qualification": "Bachelor", "status": "Active",
	})
	print("created              :", created)
	assert created["name"] == EMP and created["grade"] == "7", created
	assert created["basic_salary"] and created["basic_salary"] > 0, "no basic preview on create"
	# deprecated mirror was synced by the controller, not by the client
	assert frappe.db.get_value("Government Employee Payroll Profile", EMP, "grade_code") == "7"

	# update via the endpoint — change the stage; basic re-previewed
	updated = papi.save_employee_profile({"current_stage": 2}, name=EMP)
	print("updated stage->2     :", updated["current_stage"], "| basic:", updated["basic_salary"])
	assert updated["current_stage"] == 2, updated

	# invalid placement on update -> Arabic throw, no change persisted
	invalid = _blocked(lambda: papi.save_employee_profile({"current_stage": 99}, name=EMP))
	print("invalid update 99    : blocked =", invalid)
	assert invalid, "invalid (grade, stage) update was not rejected"
	assert frappe.db.get_value("Government Employee Payroll Profile", EMP, "current_stage") == 2

	# RBAC: a Read Only User may not create/edit through the endpoint
	def ensure_user(email, role):
		if not frappe.db.exists("User", email):
			frappe.get_doc({"doctype": "User", "email": email, "send_welcome_email": 0,
							"first_name": email.split("@")[0]}).insert(ignore_permissions=True)
		u = frappe.get_doc("User", email)
		if role not in {r.role for r in u.roles}:
			u.append("roles", {"role": role})
			u.save(ignore_permissions=True)
		return email

	ro = ensure_user("m5_readonly@test.local", "Read Only User")
	frappe.db.commit()
	frappe.set_user(ro)
	denied = _blocked(lambda: papi.save_employee_profile({
		"employee_number": "GRD-API-RO", "employee_name": "Nope",
		"rule_set": "IRAQ-2015", "grade": "7", "current_stage": 1, "status": "Active"}))
	frappe.set_user("Administrator")
	print("read-only create     : denied =", denied)
	assert denied, "Read Only User was ALLOWED to create a profile"
	assert not frappe.db.exists("Government Employee Payroll Profile", "GRD-API-RO")

	# cleanup (re-runnable)
	frappe.delete_doc("Government Employee Payroll Profile", EMP, force=True)
	frappe.db.commit()
	print("\nEMPLOYEE PROFILE API SMOKE TEST PASSED")


def demo():
	"""Phase 5 M4 — seed the demo dataset (idempotent) and verify the dashboard /
	reports / exports return meaningful data. This check COMMITS (the demo data is
	meant to persist for demonstration); re-running is safe (the seed is idempotent).
	"""
	from iraqi_government_payroll.demo import seed
	from iraqi_government_payroll.api import reports_api as rep
	from iraqi_government_payroll.api import accounting_api as acc

	summary = seed.seed_demo()
	print("seed summary        :", summary)

	# Demo structure + employees
	assert summary["entities"] == 4, summary
	emp_total = frappe.db.count("Government Employee Payroll Profile",
								{"employee_number": ["like", "DEMO-EMP-%"]})
	assert emp_total >= 25, f"expected >=25 demo employees, got {emp_total}"

	# Run states: active (Calculated) / completed (Submitted) / locked (Locked)
	states = summary["runs"]["states"]
	assert states["active"] == "Calculated", states
	assert states["completed"] == "Submitted", states
	assert states["locked"] == "Locked", states
	active, locked = summary["runs"]["active"], summary["runs"]["locked"]

	# Reports return non-empty data (active = live slips, locked = snapshots)
	assert len(rep.employee_register(active)["rows"]) == emp_total, "employee register empty"
	assert len(rep.tax_register(active)["rows"]) == emp_total, "tax register empty"
	bt = rep.bank_transfer(active)
	assert len(bt["rows"]) == emp_total and bt["incomplete_count"] > 0, "bank transfer wrong"
	assert len(rep.employee_register(locked)["rows"]) == emp_total, "locked snapshot report empty"

	# Accounting journal (proposal) balances; pension register has rows
	j = acc.journal_export(active)
	assert j["balanced"] and j["total_debit"] == j["total_credit"] and j["total_debit"] > 0, j
	pr = rep.pension_register("2024-01-01", "2024-12-31", None)
	assert pr["count"] >= 5, f"pension register too small: {pr['count']}"

	print("verified            : %d employees, reports non-empty, journal balanced, pension %d"
		  % (emp_total, pr["count"]))
	frappe.db.commit()                          # keep the demo data
	print("\nDEMO DATA SMOKE TEST PASSED")
