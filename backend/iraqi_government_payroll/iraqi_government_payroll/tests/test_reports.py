# Copyright (c) 2026, Iraqi Government Payroll
"""Tests for Phase 4 M10 payroll reports (pure + fake-frappe, no bench).

Run:  python3 -m unittest test_reports -v
"""

import os
import sys
import types
import unittest

HERE = os.path.dirname(os.path.abspath(__file__))
APP_ROOT = os.path.abspath(os.path.join(HERE, "..", ".."))
sys.path.insert(0, APP_ROOT)

from iraqi_government_payroll.services.reports import report_service as rs  # noqa: E402


def _row(emp, name, basic, allowances_lines, deduction_lines):
	"""Build a normalized row; total_earnings = basic + Σ earnings, total_deductions = Σ deductions."""
	earn = sum(a["amount"] for a in allowances_lines)
	ded = sum(d["amount"] for d in deduction_lines)
	lines = (
		[{"line_type": rs.EARNING, **a} for a in allowances_lines]
		+ [{"line_type": rs.DEDUCTION, **d} for d in deduction_lines]
	)
	return {
		"employee_profile": emp, "employee_name": name, "grade_code": "7", "stage": 1,
		"basic": basic, "total_earnings": basic + earn,
		"total_deductions": ded, "net": basic + earn - ded, "lines": lines,
	}


# Two employees with realistic components.
ROWS = [
	_row("E1", "Ali", 296000,
		 [{"component_code": "ALLOW_CERT", "component_name": "Certificate", "amount": 133200,
		   "basis_amount": 296000, "rate": 45}],
		 [{"component_code": "INCOME_TAX", "component_name": "Income Tax", "amount": 57713,
		   "basis_amount": 429200, "rate": None}]),
	_row("E2", "Sara", 200000,
		 [{"component_code": "ALLOW_POSITION", "component_name": "Position", "amount": 40000,
		   "basis_amount": 200000, "rate": 20}],
		 [{"component_code": "DED_PENSION", "component_name": "Pension", "amount": 24000,
		   "basis_amount": 200000, "rate": 12},
		  {"component_code": "INCOME_TAX", "component_name": "Income Tax", "amount": 18000,
		   "basis_amount": 240000, "rate": None}]),
]


class TestRunSummary(unittest.TestCase):
	def test_totals(self):
		s = rs.run_summary(ROWS)
		self.assertEqual(s["employees"], 2)
		self.assertEqual(s["total_basic"], 496000)
		self.assertEqual(s["total_earnings"], 296000 + 133200 + 200000 + 40000)
		self.assertEqual(s["total_deductions"], 57713 + 24000 + 18000)
		self.assertEqual(s["total_net"], (429200 - 57713) + (240000 - 42000))

	def test_empty(self):
		self.assertEqual(rs.run_summary([]),
						 {"employees": 0, "total_basic": 0, "total_earnings": 0,
						  "total_deductions": 0, "total_net": 0})


class TestEmployeeRegister(unittest.TestCase):
	def test_rows_and_totals(self):
		reg = rs.employee_register(ROWS)
		self.assertEqual(len(reg["rows"]), 2)
		ali = reg["rows"][0]
		self.assertEqual(ali["basic"], 296000)
		self.assertEqual(ali["allowances"], 133200)         # earnings - basic
		self.assertEqual(ali["deductions"], 57713)
		self.assertEqual(ali["net"], 371487)
		self.assertEqual(reg["totals"]["net"], 371487 + 198000)


class TestComponentRegisters(unittest.TestCase):
	def test_allowances(self):
		reg = rs.allowances_register(ROWS)
		self.assertEqual(reg["totals_by_component"], {"ALLOW_CERT": 133200, "ALLOW_POSITION": 40000})
		self.assertEqual(reg["grand_total"], 173200)
		# every row is an Earning component
		self.assertTrue(all(r["component_code"].startswith("ALLOW_") for r in reg["rows"]))

	def test_deductions_include_active_pension(self):
		reg = rs.deductions_register(ROWS)
		self.assertEqual(reg["totals_by_component"]["DED_PENSION"], 24000)
		self.assertEqual(reg["totals_by_component"]["INCOME_TAX"], 75713)
		self.assertEqual(reg["grand_total"], 99713)

	def test_tax_register(self):
		reg = rs.tax_register(ROWS)
		self.assertEqual(len(reg["rows"]), 2)
		self.assertEqual(reg["total_tax"], 75713)
		self.assertEqual(reg["rows"][0]["taxable"], 429200)


class TestReconciliation(unittest.TestCase):
	"""The core 'no duplicated calc' guard: every register ties back to the slips."""

	def test_registers_reconcile_to_slip_totals(self):
		summary = rs.run_summary(ROWS)
		emp = rs.employee_register(ROWS)
		allow = rs.allowances_register(ROWS)
		ded = rs.deductions_register(ROWS)
		tax = rs.tax_register(ROWS)
		# allowances grand total == Σ (earnings - basic)
		self.assertEqual(allow["grand_total"], emp["totals"]["allowances"])
		# deductions grand total == Σ deductions == summary deductions
		self.assertEqual(ded["grand_total"], emp["totals"]["deductions"])
		self.assertEqual(ded["grand_total"], summary["total_deductions"])
		# tax is a subset of deductions
		self.assertLessEqual(tax["total_tax"], ded["grand_total"])
		# net == earnings - deductions
		self.assertEqual(summary["total_net"],
						 summary["total_earnings"] - summary["total_deductions"])


# --------------------------- API source selection (fake frappe) --------------------------- #
def _raise(msg):
	raise RuntimeError(msg)


class TestReportSourceSelection(unittest.TestCase):
	"""Locked run -> Snapshot; otherwise -> Salary Slip."""

	def _load_api(self, *, workflow_state):
		queried = []
		frappe = types.ModuleType("frappe")
		frappe.throw = lambda msg, *a, **k: _raise(msg)
		frappe._ = lambda s, *a, **k: s
		frappe.db = types.SimpleNamespace(
			get_value=lambda dt, name, field: workflow_state)

		def get_all(dt, filters=None, fields=None, **k):
			queried.append(dt)
			if dt == "Salary Slip":
				return [{"name": "SLIP-1", "employee_profile": "E1", "employee_name": "Ali",
						 "grade_code": "7", "stage": 1, "basic_salary": 296000,
						 "total_earnings": 429200, "total_deductions": 57713, "net_salary": 371487}]
			if dt == "Salary Slip Line":
				return [{"parent": "SLIP-1", "line_type": "Deduction",
						 "component_code": "INCOME_TAX", "component_name": "Income Tax",
						 "amount": 57713, "basis_amount": 429200, "rate": None}]
			if dt == "Payroll Calculation Snapshot":
				return [{"name": "SNAP-1", "employee_profile": "E1", "employee_name": "Ali",
						 "grade_code": "7", "stage": 1, "gross_amount": 429200,
						 "total_deductions": 57713, "net_amount": 371487}]
			if dt == "Payroll Calculation Snapshot Line":
				return [{"parent": "SNAP-1", "line_type": "Earning",
						 "component_code": "ALLOW_CERT", "component_name": "Cert",
						 "amount": 133200, "basis_amount": 296000, "rate": 45}]
			return []
		frappe.get_all = get_all

		def whitelist(*a, **k):
			def deco(f):
				return f
			return deco if not (a and callable(a[0])) else a[0]
		frappe.whitelist = whitelist
		sys.modules["frappe"] = frappe
		import importlib
		mod = importlib.import_module("iraqi_government_payroll.api.reports_api")
		importlib.reload(mod)
		return mod, queried

	def test_active_run_reads_salary_slip(self):
		api, queried = self._load_api(workflow_state="Calculated")
		res = api.run_summary("RUN-1")
		self.assertIn("Salary Slip", queried)
		self.assertNotIn("Payroll Calculation Snapshot", queried)
		self.assertEqual(res["total_net"], 371487)

	def test_locked_run_reads_snapshot(self):
		api, queried = self._load_api(workflow_state="Locked")
		res = api.run_summary("RUN-1")
		self.assertIn("Payroll Calculation Snapshot", queried)
		self.assertEqual(res["employees"], 1)
		# basic recovered from gross - earnings (429200 - 133200)
		emp = api.employee_register("RUN-1")
		self.assertEqual(emp["rows"][0]["basic"], 296000)
		self.assertEqual(emp["rows"][0]["allowances"], 133200)


def _pension_row(emp, name, approved, cert, col, tax, other, eos, **extra):
	gross = approved + cert + col
	net = gross - tax - other
	row = {
		"employee_profile": emp, "employee_name": name, "qualification": "Bachelor",
		"service_years": 30, "average_36_months": 1010000,
		"approved_pension": approved, "certificate_allowance": cert, "cost_of_living": col,
		"gross_pension": gross, "monthly_tax": tax, "other_deductions": other,
		"net_pension": net, "end_of_service_bonus": eos,
		"status": "Approved", "calculation_date": "2026-06-01",
	}
	row.update(extra)
	return row


PENSION_ROWS = [
	_pension_row("R1", "Mahmoud", 909000, 90900, 227250, 60000, 0, 12360000),
	_pension_row("R2", "Layla", 600000, 0, 150000, 30000, 5000, 0),
]


class TestPensionRegister(unittest.TestCase):
	def test_totals(self):
		reg = rs.pension_register(PENSION_ROWS)
		self.assertEqual(reg["count"], 2)
		self.assertEqual(reg["totals"]["approved_pension"], 1509000)
		self.assertEqual(reg["totals"]["gross_pension"], 1227150 + 750000)
		self.assertEqual(reg["totals"]["net_pension"], 1167150 + 715000)
		self.assertEqual(reg["totals"]["end_of_service_bonus"], 12360000)

	def test_rows_passthrough_all_fields(self):
		reg = rs.pension_register(PENSION_ROWS)
		row = reg["rows"][0]
		for f in ("employee_profile", "employee_name", "qualification", "service_years",
				  "average_36_months", "approved_pension", "certificate_allowance",
				  "cost_of_living", "gross_pension", "monthly_tax", "other_deductions",
				  "net_pension", "end_of_service_bonus", "status", "calculation_date"):
			self.assertIn(f, row)

	def test_reconciliation(self):
		# gross == approved + cert + col ; net == gross - tax - other
		for r in PENSION_ROWS:
			self.assertEqual(r["gross_pension"],
							 r["approved_pension"] + r["certificate_allowance"] + r["cost_of_living"])
			self.assertEqual(r["net_pension"],
							 r["gross_pension"] - r["monthly_tax"] - r["other_deductions"])


class TestPensionSourceSelection(unittest.TestCase):
	"""Finalized (Approved) record with a snapshot -> snapshot figures; else -> live."""

	def _load_api(self, *, record, snapshot=None):
		frappe = types.ModuleType("frappe")
		frappe.throw = lambda msg, *a, **k: _raise(msg)
		frappe._ = lambda s, *a, **k: s
		frappe.db = types.SimpleNamespace(
			get_value=lambda dt, name, field: "Bachelor")  # qualification lookup

		def get_all(dt, filters=None, fields=None, **k):
			if dt == "Pension Calculation":
				return [record]
			if dt == "Payroll Calculation Snapshot":
				return [snapshot] if snapshot else []
			return []
		frappe.get_all = get_all

		def whitelist(*a, **k):
			def deco(f):
				return f
			return deco if not (a and callable(a[0])) else a[0]
		frappe.whitelist = whitelist
		sys.modules["frappe"] = frappe
		import importlib
		mod = importlib.import_module("iraqi_government_payroll.api.reports_api")
		importlib.reload(mod)
		return mod

	def test_finalized_reads_snapshot(self):
		# live record says net 999; snapshot output says net 715000 -> snapshot wins
		record = {"name": "PC1", "employee_profile": "R2", "employee_name": "Layla",
				  "calculation_date": "2026-06-01", "status": "Approved",
				  "period_date": "2026-06-01", "service_years": 99, "net_pension": 999,
				  "approved_pension": 1, "certificate_allowance": 1, "cost_of_living": 1,
				  "gross_pension": 3, "monthly_tax": 1, "other_deductions": 1,
				  "average_36_months": 1, "end_of_service_bonus": 1}
		out_json = ('{"service_months": 360, "avg36": 1010000, "approved_pension": 600000,'
					' "certificate_allowance": 0, "cost_of_living": 150000,'
					' "gross_pension": 750000, "monthly_tax": 30000, "other_deductions": 5000,'
					' "net_pension": 715000, "end_of_service_bonus": 0}')
		snap = {"name": "SN1", "employee_name": "Layla", "gross_amount": 750000,
				"net_amount": 715000, "output_snapshot": out_json}
		api = self._load_api(record=record, snapshot=snap)
		reg = api.pension_register("2026-06-01", "2026-06-30", "Approved")
		row = reg["rows"][0]
		self.assertEqual(row["net_pension"], 715000)        # from snapshot, not 999
		self.assertEqual(row["service_years"], 30)          # 360 months / 12
		self.assertEqual(row["status"], "Approved")
		self.assertEqual(row["qualification"], "Bachelor")  # profile lookup
		self.assertEqual(row["calculation_date"], "2026-06-01")

	def test_non_finalized_reads_live(self):
		record = {"name": "PC2", "employee_profile": "R3", "employee_name": "Omar",
				  "calculation_date": "2026-06-02", "status": "Calculated",
				  "period_date": "2026-06-01", "service_years": 28, "net_pension": 500000,
				  "approved_pension": 400000, "certificate_allowance": 0, "cost_of_living": 100000,
				  "gross_pension": 500000, "monthly_tax": 0, "other_deductions": 0,
				  "average_36_months": 800000, "end_of_service_bonus": 0}
		# snapshot present but must be ignored (status not finalized)
		api = self._load_api(record=record, snapshot={"name": "X", "output_snapshot": "{}"})
		reg = api.pension_register("2026-06-01", "2026-06-30", None)
		row = reg["rows"][0]
		self.assertEqual(row["net_pension"], 500000)        # from live record
		self.assertEqual(row["service_years"], 28)
		self.assertEqual(row["status"], "Calculated")


class TestBankTransfer(unittest.TestCase):
	"""Pure bank_transfer flagging: complete = (iban OR bank_account) AND net>0."""

	def _rows(self):
		return [
			# complete: has account + net
			{"employee_profile": "E1", "employee_name": "Ali", "net": 371487,
			 "iban": "", "bank_name": "Rafidain", "bank_account": "12345", "national_id": "X"},
			# complete: has iban only + net
			{"employee_profile": "E2", "employee_name": "Sara", "net": 200000,
			 "iban": "IQ98...", "bank_name": "", "bank_account": "", "national_id": ""},
			# incomplete: no account, no iban
			{"employee_profile": "E3", "employee_name": "Omar", "net": 150000,
			 "iban": "", "bank_name": "", "bank_account": "", "national_id": ""},
			# incomplete: has account but net 0
			{"employee_profile": "E4", "employee_name": "Hana", "net": 0,
			 "iban": "", "bank_name": "", "bank_account": "999", "national_id": ""},
		]

	def test_all_rows_included_and_flagged(self):
		res = rs.bank_transfer(self._rows())
		self.assertEqual(res["count"], 4)               # nothing skipped
		flags = {r["employee_profile"]: r["bank_complete"] for r in res["rows"]}
		self.assertEqual(flags, {"E1": True, "E2": True, "E3": False, "E4": False})
		self.assertEqual(res["incomplete_count"], 2)

	def test_missing_reasons(self):
		res = rs.bank_transfer(self._rows())
		by = {r["employee_profile"]: r["missing"] for r in res["rows"]}
		self.assertEqual(by["E1"], [])
		self.assertEqual(by["E3"], ["bank_account"])    # no payable identifier
		self.assertEqual(by["E4"], ["net"])             # has account, net 0

	def test_total_net_reconciles(self):
		res = rs.bank_transfer(self._rows())
		# total_net is the plain sum of the slip-sourced net — ties to the register
		self.assertEqual(res["total_net"], 371487 + 200000 + 150000 + 0)


class TestBankTransferApi(unittest.TestCase):
	"""Net from Salary Slip (active) vs Snapshot (locked); bank from profile."""

	def _load_api(self, *, workflow_state):
		frappe = types.ModuleType("frappe")
		frappe.throw = lambda msg, *a, **k: _raise(msg)
		frappe._ = lambda s, *a, **k: s
		frappe.db = types.SimpleNamespace(get_value=lambda dt, name, field: workflow_state)

		def get_all(dt, filters=None, fields=None, **k):
			if dt == "Salary Slip" and fields and "basic_salary" in fields:
				return [{"name": "SLIP-1", "employee_profile": "E1", "employee_name": "Ali",
						 "grade_code": "7", "stage": 1, "basic_salary": 296000,
						 "total_earnings": 429200, "total_deductions": 57713, "net_salary": 371487}]
			if dt == "Salary Slip":            # name-only lookup (snapshot path)
				return [{"name": "SLIP-1"}]
			if dt == "Salary Slip Line":
				return []
			if dt == "Payroll Calculation Snapshot":
				return [{"name": "SNAP-1", "employee_profile": "E1", "employee_name": "Ali",
						 "grade_code": "7", "stage": 1, "gross_amount": 429200,
						 "total_deductions": 57713, "net_amount": 999999}]
			if dt == "Payroll Calculation Snapshot Line":
				return []
			if dt == "Government Employee Payroll Profile":
				return [{"name": "E1", "iban": "", "bank_name": "Rafidain",
						 "bank_account": "12345", "national_id": "N1"}]
			return []
		frappe.get_all = get_all

		def whitelist(*a, **k):
			def deco(f):
				return f
			return deco if not (a and callable(a[0])) else a[0]
		frappe.whitelist = whitelist
		sys.modules["frappe"] = frappe
		import importlib
		mod = importlib.import_module("iraqi_government_payroll.api.reports_api")
		importlib.reload(mod)
		return mod

	def test_active_uses_salary_slip_net_and_profile_bank(self):
		api = self._load_api(workflow_state="Calculated")
		res = api.bank_transfer("RUN-1")
		row = res["rows"][0]
		self.assertEqual(row["net"], 371487)            # live slip net
		self.assertEqual(row["bank_account"], "12345")  # joined from profile
		self.assertTrue(row["bank_complete"])

	def test_locked_uses_snapshot_net(self):
		api = self._load_api(workflow_state="Locked")
		res = api.bank_transfer("RUN-1")
		self.assertEqual(res["rows"][0]["net"], 999999)  # immutable snapshot net


class TestColumnSpecs(unittest.TestCase):
	def test_all_seven_reports_have_specs(self):
		from iraqi_government_payroll.services.reports import report_columns as rc
		expected = {"run_summary", "employee_register", "allowances_register",
					"deductions_register", "tax_register", "pension_register",
					"bank_transfer"}
		self.assertEqual(set(rc.REPORT_SPECS), expected)
		for spec in rc.REPORT_SPECS.values():
			self.assertTrue(spec["columns"])               # non-empty
			self.assertTrue(callable(spec["rows"]))
			self.assertTrue(callable(spec["totals"]))


class TestXlsxExport(unittest.TestCase):
	"""Build a workbook then REOPEN it and assert structure + cells."""

	def _reopen(self, content):
		import io
		from openpyxl import load_workbook
		return load_workbook(io.BytesIO(content))

	def test_headers_rows_totals_and_rtl(self):
		from iraqi_government_payroll.services.reports import xlsx_export
		columns = [("employee_name", "الاسم"), ("net", "الصافي")]
		rows = [{"employee_name": "Ali", "net": 371487}, {"employee_name": "Sara", "net": 200000}]
		content = xlsx_export.build_workbook("كشف", columns, rows, totals={"net": 571487})
		self.assertTrue(content[:2] == b"PK")              # xlsx is a zip
		ws = self._reopen(content).active
		self.assertTrue(ws.sheet_view.rightToLeft)         # RTL sheet view
		self.assertEqual([c.value for c in ws[1]], ["الاسم", "الصافي"])  # header
		self.assertEqual(ws.cell(row=2, column=1).value, "Ali")
		self.assertEqual(ws.cell(row=2, column=2).value, 371487)        # numeric cell
		self.assertEqual(ws.max_row, 4)                    # header + 2 rows + totals
		self.assertEqual(ws.cell(row=4, column=1).value, "الإجمالي")    # totals label
		self.assertEqual(ws.cell(row=4, column=2).value, 571487)
		self.assertTrue(ws.cell(row=1, column=1).font.bold)            # bold header

	def test_bool_and_list_cells(self):
		from iraqi_government_payroll.services.reports import xlsx_export
		columns = [("bank_complete", "مكتمل"), ("missing", "النواقص")]
		rows = [{"bank_complete": True, "missing": []},
				{"bank_complete": False, "missing": ["bank_account", "net"]}]
		ws = self._reopen(xlsx_export.build_workbook("بنك", columns, rows)).active
		self.assertEqual(ws.cell(row=2, column=1).value, "نعم")
		self.assertEqual(ws.cell(row=3, column=1).value, "لا")
		self.assertEqual(ws.cell(row=3, column=2).value, "bank_account، net")


class TestExportDispatch(unittest.TestCase):
	"""render_report_xlsx through the real aggregator + serializer (fake frappe)."""

	def _load_api(self, *, workflow_state="Calculated"):
		frappe = types.ModuleType("frappe")
		frappe.throw = lambda msg, *a, **k: _raise(msg)
		frappe._ = lambda s, *a, **k: s
		frappe.db = types.SimpleNamespace(get_value=lambda dt, name, field: workflow_state)

		def get_all(dt, filters=None, fields=None, **k):
			if dt == "Salary Slip" and fields and "basic_salary" in fields:
				return [{"name": "S1", "employee_profile": "E1", "employee_name": "Ali",
						 "grade_code": "7", "stage": 1, "basic_salary": 296000,
						 "total_earnings": 429200, "total_deductions": 57713, "net_salary": 371487}]
			if dt == "Salary Slip Line":
				return []
			return []
		frappe.get_all = get_all

		def whitelist(*a, **k):
			def deco(f):
				return f
			return deco if not (a and callable(a[0])) else a[0]
		frappe.whitelist = whitelist
		sys.modules["frappe"] = frappe
		import importlib
		mod = importlib.import_module("iraqi_government_payroll.api.reports_api")
		importlib.reload(mod)
		return mod

	def test_employee_register_xlsx(self):
		import io
		from openpyxl import load_workbook
		api = self._load_api()
		content = api.render_report_xlsx("employee_register", run="R1")
		ws = load_workbook(io.BytesIO(content)).active
		self.assertEqual(ws.cell(row=1, column=1).value, "الموظف")   # header
		self.assertEqual(ws.cell(row=2, column=2).value, "Ali")
		# last row is the bold totals row with net total
		self.assertEqual(ws.cell(row=ws.max_row, column=1).value, "الإجمالي")

	def test_unknown_report_and_format_rejected(self):
		api = self._load_api()
		with self.assertRaises(Exception):
			api.render_report_xlsx("nope", run="R1")
		with self.assertRaises(Exception):
			api.export_report("employee_register", fmt="pdf", run="R1")


class TestPdfHtml(unittest.TestCase):
	"""Pure build_html assertions — no frappe, no wkhtmltopdf."""

	def _pdf(self):
		from iraqi_government_payroll.services.reports import pdf_export
		return pdf_export

	def test_structure_title_rtl_font_rows_totals(self):
		cols = [("employee_name", "الاسم"), ("net", "الصافي")]
		rows = [{"employee_name": "علي", "net": 371487},
				{"employee_name": "سارة", "net": 200000}]
		html = self._pdf().build_html("كشف الرواتب", cols, rows, totals={"net": 571487})
		self.assertIsInstance(html, str)
		self.assertIn("<!DOCTYPE html", html)
		self.assertIn('dir="rtl"', html)
		self.assertIn("@font-face", html)              # bundled Amiri embedded
		self.assertIn("base64,", html)
		self.assertIn("كشف الرواتب", html)             # report title
		self.assertIn("الاسم", html)                   # Arabic header label
		self.assertIn("الصافي", html)
		self.assertIn("الإجمالي", html)                # totals row label
		self.assertEqual(html.count("<tr>"), 4)        # header + 2 data + totals
		self.assertIn("371487", html)

	def test_empty_rows_still_valid_html(self):
		html = self._pdf().build_html("فارغ", [("employee_name", "الاسم")], [])
		self.assertIn("<!DOCTYPE html", html)
		self.assertIn("الاسم", html)                   # header still rendered
		self.assertEqual(html.count("<tr>"), 1)        # header only, no data rows
		self.assertTrue(html.rstrip().endswith("</html>"))

	def test_escaping_bool_and_list(self):
		cols = [("name", "الاسم"), ("complete", "مكتمل"), ("missing", "النواقص")]
		rows = [{"name": "A & B <x>", "complete": False,
				 "missing": ["bank_account", "net"]}]
		html = self._pdf().build_html("t", cols, rows)
		self.assertIn("A &amp; B &lt;x&gt;", html)     # HTML-escaped (render-safe)
		self.assertIn("لا", html)                      # bool False -> لا
		self.assertIn("bank_account، net", html)       # list joined, incomplete kept


class TestPdfDispatch(unittest.TestCase):
	"""render_report_html / render_report_pdf through the real aggregator+spec,
	with frappe (and wkhtmltopdf) faked."""

	def _load_api(self, *, pdf_bytes=b"%PDF-1.4 fake-bytes", workflow_state="Calculated"):
		frappe = types.ModuleType("frappe")
		frappe.throw = lambda msg, *a, **k: _raise(msg)
		frappe._ = lambda s, *a, **k: s
		frappe.response = {}
		frappe.db = types.SimpleNamespace(get_value=lambda dt, name, field: workflow_state)

		def get_all(dt, filters=None, fields=None, **k):
			if dt == "Salary Slip" and fields and "basic_salary" in fields:
				return [{"name": "S1", "employee_profile": "E1", "employee_name": "علي",
						 "grade_code": "7", "stage": 1, "basic_salary": 296000,
						 "total_earnings": 429200, "total_deductions": 57713, "net_salary": 371487}]
			return []
		frappe.get_all = get_all

		# fake frappe.utils.pdf.get_pdf (no wkhtmltopdf on the host)
		utils = types.ModuleType("frappe.utils")
		pdf_mod = types.ModuleType("frappe.utils.pdf")
		pdf_mod.get_pdf = lambda html, options=None: pdf_bytes
		utils.pdf = pdf_mod
		frappe.utils = utils

		def whitelist(*a, **k):
			def deco(f):
				return f
			return deco if not (a and callable(a[0])) else a[0]
		frappe.whitelist = whitelist
		sys.modules["frappe"] = frappe
		sys.modules["frappe.utils"] = utils
		sys.modules["frappe.utils.pdf"] = pdf_mod
		import importlib
		mod = importlib.import_module("iraqi_government_payroll.api.reports_api")
		importlib.reload(mod)
		return mod, frappe

	def test_html_uses_shared_report_columns(self):
		from iraqi_government_payroll.services.reports import report_columns as rc
		api, _ = self._load_api()
		html = api.render_report_html("employee_register", run="R1")
		for _key, header in rc.REPORT_SPECS["employee_register"]["columns"]:
			self.assertIn(header, html)                # every shared header present
		self.assertIn(rc.REPORT_SPECS["employee_register"]["title"], html)

	def test_pdf_returns_bytes(self):
		api, _ = self._load_api()
		pdf = api.render_report_pdf("employee_register", run="R1")
		self.assertTrue(pdf.startswith(b"%PDF"))

	def test_export_report_pdf_sets_response(self):
		api, frappe = self._load_api()
		api.export_report("employee_register", fmt="pdf", run="R1")
		self.assertEqual(frappe.response["filename"], "employee_register.pdf")
		self.assertTrue(frappe.response["filecontent"].startswith(b"%PDF"))

	def test_xlsx_still_works_and_unknown_format_rejected(self):
		api, frappe = self._load_api()
		api.export_report("employee_register", fmt="xlsx", run="R1")
		self.assertEqual(frappe.response["filename"], "employee_register.xlsx")
		self.assertTrue(frappe.response["filecontent"][:2] == b"PK")   # xlsx unchanged
		with self.assertRaises(Exception):
			api.export_report("employee_register", fmt="csv", run="R1")


if __name__ == "__main__":
	unittest.main()
